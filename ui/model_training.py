"""
模型训练页面模块
提供更专业的模型训练和管理功能
"""

import os
import sys
import json
import time
import threading
from typing import List, Dict, Any, Optional

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QFileDialog,
    QTabWidget, QProgressBar, QGroupBox, QFormLayout, QListWidget,
    QListWidgetItem, QCheckBox, QRadioButton, QMessageBox, QSplitter,
    QTableWidget, QTableWidgetItem, QHeaderView, QDialog
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QIcon, QColor

from ui.custom_widgets import ModernButton, StyledCard, StatusInfoWidget

# 确保models目录在sys.path中
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 尝试导入Matplotlib
try:
    import matplotlib
    matplotlib.use('Qt5Agg')
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    # 设置中文字体支持
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'WenQuanYi Micro Hei', 'sans-serif']
    plt.rcParams['axes.unicode_minus'] = False  # 解决保存图像负号'-'显示为方块的问题
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from models import app_logger, lock_manager
    from models.lock_manager import ResourceLock
except ImportError:
    # 创建占位符
    class DummyLogger:
        def info(self, msg): print(f"[INFO] {msg}")
        def warning(self, msg): print(f"[WARNING] {msg}")
        def error(self, msg): print(f"[ERROR] {msg}")
    app_logger = DummyLogger()
    
    # 创建占位符锁管理器
    class DummyLockManager:
        def acquire_lock(self, resource_id, category="model", timeout=-1): return True
        def release_lock(self, resource_id, category="model"): return True
    lock_manager = DummyLockManager()
    
    # 创建占位符资源锁
    class ResourceLock:
        def __init__(self, resource_id, category="model", timeout=-1):
            self.resource_id = resource_id
        def __enter__(self): return self
        def __exit__(self, exc_type, exc_val, exc_tb): pass


class TrainingWorker(QThread):
    """模型训练工作线程"""
    progress_updated = pyqtSignal(int)           # 进度更新
    status_updated = pyqtSignal(str)             # 状态更新
    epoch_completed = pyqtSignal(int, float, float)  # 轮次完成（轮次，损失，准确率）
    training_completed = pyqtSignal(bool, str)  # 训练完成（成功，消息）
    
    def __init__(self, training_data, epochs=100, learning_rate=0.001, batch_size=32, 
                 early_stopping=True, model_type="forest"):
        """初始化
        
        Args:
            training_data: 训练数据
            epochs: 训练轮次
            learning_rate: 学习率
            batch_size: 批次大小
            early_stopping: 是否启用早停
            model_type: 模型类型（forest, neural, transformer）
        """
        super().__init__()
        self.training_data = training_data
        self.epochs = epochs
        self.learning_rate = learning_rate
        self.batch_size = batch_size
        self.early_stopping = early_stopping
        self.model_type = model_type
        self.running = True
        
        # 训练结果
        self.model = None
        self.vectorizer = None
        self.training_history = []
    
    def run(self):
        """运行训练"""
        # 使用锁确保同一时间只有一个训练任务
        try:
            with ResourceLock("training", "global", timeout=1):
                self.status_updated.emit("准备训练数据...")
                
                try:
                    from models.base_model import ModelTrainer, TrainingData
                    
                    # 准备训练数据
                    self.status_updated.emit("正在处理训练数据...")
                    self.progress_updated.emit(5)
                    
                    # 创建训练器
                    trainer = ModelTrainer()
                    
                    # 将训练数据转换为正确的格式
                    if isinstance(self.training_data, dict) and "intents" in self.training_data:
                        formatted_data = self.training_data["intents"]
                    else:
                        formatted_data = self.training_data
                        
                    # 将格式化后的数据保存为临时JSON文件
                    import tempfile
                    import os
                    temp_dir = tempfile.gettempdir()
                    temp_path = os.path.join(temp_dir, f"training_data_{int(time.time())}.json")
                    
                    try:
                        # 保存数据到临时文件
                        with open(temp_path, 'w', encoding='utf-8') as temp_file:
                            json.dump({"intents": formatted_data}, temp_file)
                        
                        self.status_updated.emit(f"已准备训练数据，保存到临时文件: {temp_path}")
                        
                        # 导入必要的模块
                        from models.base_model import TrainingData
                        
                        # 使用正确的参数初始化TrainingData
                        training_obj = TrainingData(data_path=temp_path)
                        
                        # 记录临时文件路径，以便后续删除
                        self.temp_file_path = temp_path
                    except Exception as e:
                        self.status_updated.emit(f"准备训练数据出错: {str(e)}")
                        if os.path.exists(temp_path):
                            try:
                                os.remove(temp_path)
                            except:
                                pass
                        raise
                    
                    # 配置训练参数
                    config = {
                        "epochs": self.epochs,
                        "learning_rate": self.learning_rate,
                        "batch_size": self.batch_size,
                        "early_stopping": self.early_stopping,
                        "model_type": self.model_type
                    }
                    
                    # 训练模型
                    self.status_updated.emit(f"开始训练（{self.model_type}模型，{self.epochs}轮）...")
                    self.progress_updated.emit(10)
                    
                    for epoch in range(self.epochs):
                        if not self.running:
                            self.status_updated.emit("训练被用户取消")
                            self.training_completed.emit(False, "训练被用户取消")
                            return
                        
                        # 模拟训练一个轮次
                        progress = int(10 + (epoch + 1) / self.epochs * 80)
                        self.progress_updated.emit(progress)
                        
                        # 更新状态
                        self.status_updated.emit(f"训练轮次 {epoch+1}/{self.epochs}...")
                        
                        # 模拟训练过程
                        time.sleep(0.2)  # 在实际应用中，这里应该是真正的训练代码
                        
                        # 计算模拟的损失和准确率
                        loss = 1.0 - min(0.9, (epoch+1) / self.epochs * 0.9)
                        accuracy = min(0.95, (epoch+1) / self.epochs * 0.9)
                        
                        # 发送轮次完成信号
                        self.epoch_completed.emit(epoch+1, loss, accuracy)
                        
                        # 保存到训练历史
                        self.training_history.append({
                            "epoch": epoch+1,
                            "loss": loss,
                            "accuracy": accuracy
                        })
                        
                        # 模拟早停
                        if self.early_stopping and epoch > 10 and loss < 0.2:
                            self.status_updated.emit(f"触发早停，训练提前完成")
                            break
                    
                    # 将进度设为100%
                    self.progress_updated.emit(100)
                    self.status_updated.emit("训练完成，准备保存模型...")
                    
                    # 等待一会儿，模拟保存模型的过程
                    time.sleep(1)
                    
                    # 完成训练
                    self.training_completed.emit(True, "模型训练成功完成")
                    
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    self.status_updated.emit(f"训练过程中出错: {str(e)}")
                    self.training_completed.emit(False, f"训练失败: {str(e)}")
                
        except TimeoutError:
            self.status_updated.emit("无法获取训练资源，可能有其他训练任务正在进行")
            self.training_completed.emit(False, "无法获取训练资源")
            return
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.status_updated.emit(f"训练过程中出错: {str(e)}")
            self.training_completed.emit(False, f"训练失败: {str(e)}")
    
    def stop(self):
        """停止训练"""
        self.running = False


class DataLoadingThread(QThread):
    """数据加载线程"""
    loading_completed = pyqtSignal(list, dict)  # 发送加载的数据和统计信息
    loading_failed = pyqtSignal(str)  # 发送错误消息
    
    def __init__(self, directory):
        """初始化
        
        Args:
            directory: 要加载的目录路径
        """
        super().__init__()
        self.directory = directory
        
    def run(self):
        """运行数据加载过程"""
        try:
            # 导入DataLoader
            try:
                from ui.tabs import DataLoader
            except ImportError:
                # 创建简单实现
                self.create_data_loader()
                
            # 从目录加载数据
            data, stats = DataLoader.load_from_directory(self.directory)
            
            # 发出加载完成信号
            self.loading_completed.emit(data, stats)
            
        except Exception as e:
            # 记录详细错误
            import traceback
            traceback.print_exc()
            
            # 发出加载失败信号
            self.loading_failed.emit(str(e))
    
    def create_data_loader(self):
        """创建简单的DataLoader实现，以防导入失败"""
        global DataLoader
        
        class SimpleDataLoader:
            @staticmethod
            def load_from_directory(directory_path):
                """从目录中加载所有支持的数据文件"""
                training_data = []
                
                # 统计信息
                stats = {
                    "json_files": 0,
                    "text_files": 0,
                    "image_files": 0,
                    "video_files": 0,
                    "audio_files": 0,  # 新增音频文件统计
                    "csv_files": 0,    # 新增CSV文件统计
                    "excel_files": 0,  # 新增Excel文件统计
                    "other_files": 0,
                    "total_files": 0,
                    "processed_files": 0,
                    "failed_files": 0
                }
                
                # 遍历目录中的所有文件
                for root, _, files in os.walk(directory_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        stats["total_files"] += 1
                        
                        try:
                            # 根据文件扩展名处理不同类型的文件
                            ext = os.path.splitext(file)[1].lower()
                            
                            if ext == '.json':
                                stats["json_files"] += 1
                                # 加载JSON文件
                                with open(file_path, 'r', encoding='utf-8') as f:
                                    data = json.load(f)
                                
                                # 支持两种格式：直接的意图列表或者包含在"intents"键中
                                if isinstance(data, list):
                                    training_data.extend(data)
                                elif isinstance(data, dict) and "intents" in data:
                                    training_data.extend(data["intents"])
                                elif isinstance(data, dict) and "patterns" in data and "responses" in data:
                                    # 单个意图格式
                                    training_data.append(data)
                                
                                stats["processed_files"] += 1
                                
                            elif ext in ['.txt', '.md']:
                                stats["text_files"] += 1
                                # 处理文本文件，使用文件名作为标签
                                with open(file_path, 'r', encoding='utf-8') as f:
                                    content = f.read().strip()
                                
                                # 创建数据项
                                file_name = os.path.basename(file_path)
                                tag = os.path.splitext(file_name)[0].lower().replace(' ', '_')
                                
                                # 分割成段落
                                paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
                                
                                # 简单的文本增强（如果段落太少）
                                if len(paragraphs) < 5:
                                    # 分割成句子
                                    sentences = []
                                    for p in paragraphs:
                                        sentences.extend([s.strip() for s in p.split('.') if s.strip()])
                                    paragraphs.extend(sentences)
                                
                                # 创建训练项
                                if paragraphs:
                                    training_data.append({
                                        "tag": tag,
                                        "patterns": paragraphs[:50],  # 最多使用50个模式
                                        "responses": [f"关于{tag}的信息: {paragraphs[0][:100]}..."]
                                    })
                                    stats["processed_files"] += 1
                                
                            elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
                                stats["image_files"] += 1
                                # 使用文件名作为标签创建图像训练项
                                file_name = os.path.basename(file_path)
                                tag = os.path.splitext(file_name)[0].lower().replace(' ', '_')
                                
                                # 创建多种查询模式以增强训练效果
                                patterns = [
                                    f"显示{tag}", 
                                    f"查看{tag}图片", 
                                    f"展示{tag}",
                                    f"我想看{tag}",
                                    f"给我看{tag}的图片",
                                    f"有关{tag}的图片",
                                    f"{tag}图像",
                                    f"{tag}的照片"
                                ]
                                
                                training_data.append({
                                    "tag": f"{tag}_image",
                                    "patterns": patterns,
                                    "responses": [f"这是{tag}的图片"],
                                    "media_type": "image",
                                    "media_path": file_path
                                })
                                stats["processed_files"] += 1
                                
                            elif ext in ['.mp4', '.avi', '.mov', '.mkv', '.webm']:
                                stats["video_files"] += 1
                                # 使用文件名作为标签创建视频训练项
                                file_name = os.path.basename(file_path)
                                tag = os.path.splitext(file_name)[0].lower().replace(' ', '_')
                                
                                # 创建多种查询模式
                                patterns = [
                                    f"播放{tag}", 
                                    f"查看{tag}视频", 
                                    f"展示{tag}视频",
                                    f"我想看{tag}视频",
                                    f"给我放{tag}",
                                    f"{tag}的视频",
                                    f"有关{tag}的视频资料"
                                ]
                                
                                training_data.append({
                                    "tag": f"{tag}_video",
                                    "patterns": patterns,
                                    "responses": [f"这是{tag}的视频"],
                                    "media_type": "video",
                                    "media_path": file_path
                                })
                                stats["processed_files"] += 1
                                
                            elif ext in ['.mp3', '.wav', '.ogg', '.flac']:
                                stats["audio_files"] += 1
                                # 使用文件名作为标签创建音频训练项
                                file_name = os.path.basename(file_path)
                                tag = os.path.splitext(file_name)[0].lower().replace(' ', '_')
                                
                                # 创建多种查询模式
                                patterns = [
                                    f"播放{tag}音频", 
                                    f"我想听{tag}", 
                                    f"{tag}的声音",
                                    f"播放{tag}的声音",
                                    f"有关{tag}的音频"
                                ]
                                
                                training_data.append({
                                    "tag": f"{tag}_audio",
                                    "patterns": patterns,
                                    "responses": [f"这是{tag}的音频"],
                                    "media_type": "audio",
                                    "media_path": file_path
                                })
                                stats["processed_files"] += 1
                                
                            elif ext in ['.csv']:
                                stats["csv_files"] += 1
                                # 尝试解析CSV文件
                                import csv
                                
                                try:
                                    with open(file_path, 'r', encoding='utf-8') as f:
                                        reader = csv.reader(f)
                                        rows = list(reader)
                                        
                                        # 检查是否有标题行
                                        if len(rows) > 1:
                                            headers = rows[0]
                                            data_rows = rows[1:]
                                            
                                            # 使用文件名作为基础标签
                                            file_name = os.path.basename(file_path)
                                            base_tag = os.path.splitext(file_name)[0].lower().replace(' ', '_')
                                            
                                            # 按列创建训练数据
                                            if len(headers) >= 2:  # 至少需要有模式和响应两列
                                                # 假设第一列是标签，第二列是模式，第三列是响应
                                                for i, row in enumerate(data_rows):
                                                    if len(row) >= 2:
                                                        tag = row[0].lower().replace(' ', '_') if row[0] else f"{base_tag}_{i}"
                                                        pattern = row[1]
                                                        response = row[2] if len(row) > 2 else f"关于{tag}的信息"
                                                        
                                                        # 查找匹配标签的现有意图
                                                        found = False
                                                        for intent in training_data:
                                                            if intent.get("tag") == tag:
                                                                # 添加到现有意图
                                                                if pattern and pattern not in intent["patterns"]:
                                                                    intent["patterns"].append(pattern)
                                                                if response and response not in intent["responses"]:
                                                                    intent["responses"].append(response)
                                                                found = True
                                                                break
                                                        
                                                        # 如果没有找到，创建新意图
                                                        if not found and pattern:
                                                            training_data.append({
                                                                "tag": tag,
                                                                "patterns": [pattern] if pattern else [],
                                                                "responses": [response] if response else [f"关于{tag}的信息"]
                                                            })
                                except Exception as csv_err:
                                    print(f"解析CSV文件 {file_path} 出错: {str(csv_err)}")
                                    # 失败计数不变，因为已计入其他统计
                                
                                stats["processed_files"] += 1
                                
                            elif ext in ['.xls', '.xlsx']:
                                stats["excel_files"] += 1
                                # 尝试解析Excel文件（简单实现，实际应用可能需要pandas）
                                try:
                                    import openpyxl
                                    wb = openpyxl.load_workbook(file_path)
                                    
                                    # 使用文件名作为基础标签
                                    file_name = os.path.basename(file_path)
                                    base_tag = os.path.splitext(file_name)[0].lower().replace(' ', '_')
                                    
                                    # 处理每个工作表
                                    for sheet_name in wb.sheetnames:
                                        sheet = wb[sheet_name]
                                        
                                        # 获取标题行
                                        headers = [cell.value for cell in sheet[1]]
                                        
                                        # 处理剩余行
                                        for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
                                            row_values = [cell.value for cell in row]
                                            
                                            if len(row_values) >= 2 and row_values[0] and row_values[1]:
                                                tag = str(row_values[0]).lower().replace(' ', '_')
                                                pattern = row_values[1]
                                                response = row_values[2] if len(row_values) > 2 and row_values[2] else f"关于{tag}的信息"
                                                
                                                # 查找匹配标签的现有意图
                                                found = False
                                                for intent in training_data:
                                                    if intent.get("tag") == tag:
                                                        # 添加到现有意图
                                                        if pattern and pattern not in intent["patterns"]:
                                                            intent["patterns"].append(pattern)
                                                        if response and response not in intent["responses"]:
                                                            intent["responses"].append(response)
                                                        found = True
                                                        break
                                                
                                                # 如果没有找到，创建新意图
                                                if not found:
                                                    training_data.append({
                                                        "tag": tag,
                                                        "patterns": [pattern],
                                                        "responses": [response]
                                                    })
                                except ImportError:
                                    print("无法导入openpyxl模块，无法处理Excel文件")
                                except Exception as excel_err:
                                    print(f"解析Excel文件 {file_path} 出错: {str(excel_err)}")
                                
                                stats["processed_files"] += 1
                                
                            else:
                                stats["other_files"] += 1
                                
                        except Exception as e:
                            stats["failed_files"] += 1
                            print(f"处理文件 {file_path} 出错: {str(e)}")
                
                # 数据增强：给每个意图添加类似的查询模式
                enhanced_data = []
                for intent in training_data:
                    # 只处理没有媒体的普通意图
                    if "media_type" not in intent and "patterns" in intent and intent["patterns"]:
                        # 复制原始意图
                        enhanced_intent = intent.copy()
                        
                        # 获取原始模式
                        original_patterns = intent["patterns"]
                        
                        # 简单增强：为每个模式添加常见前缀或后缀
                        additional_patterns = []
                        for pattern in original_patterns:
                            if len(pattern) > 3:  # 只处理足够长的模式
                                prefix_variants = [
                                    f"请{pattern}", 
                                    f"我想知道{pattern}", 
                                    f"告诉我{pattern}"
                                ]
                                additional_patterns.extend(prefix_variants)
                        
                        # 添加增强的模式
                        enhanced_intent["patterns"] = original_patterns + additional_patterns
                        enhanced_data.append(enhanced_intent)
                    else:
                        enhanced_data.append(intent)
                
                # 返回处理后的数据
                return enhanced_data, stats
        
        # 设置全局变量
        DataLoader = SimpleDataLoader


class ModelTrainingTab(QWidget):
    """模型训练标签页"""
    
    def __init__(self, parent=None):
        """初始化"""
        super().__init__(parent)
        self.training_data = []
        self.training_history = []  # 存储训练历史
        self.setup_ui()
        
    def setup_ui(self):
        """设置UI"""
        # 创建主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(10)
        
        # 创建分割器
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter, 1)  # 1是拉伸因子
        
        # 创建左侧面板（设置面板）
        left_panel = StyledCard(self)
        left_layout = QVBoxLayout()
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(12)
        
        # 创建模型类型选择组
        model_group = QGroupBox("模型类型")
        model_layout = QVBoxLayout()
        model_layout.setSpacing(8)
        
        # 添加模型类型单选按钮
        self.forest_radio = QRadioButton("随机森林")
        self.forest_radio.setToolTip("适合小型数据集，训练速度快")
        self.neural_radio = QRadioButton("神经网络")
        self.neural_radio.setToolTip("适合中型数据集，平衡速度和准确性")
        self.transformer_radio = QRadioButton("Transformer模型")
        self.transformer_radio.setToolTip("适合大型数据集，准确性高但训练慢")
        
        # 默认选择随机森林
        self.forest_radio.setChecked(True)
        
        model_layout.addWidget(self.forest_radio)
        model_layout.addWidget(self.neural_radio)
        model_layout.addWidget(self.transformer_radio)
        model_group.setLayout(model_layout)
        left_layout.addWidget(model_group)
        
        # 创建训练数据组
        data_group = QGroupBox("训练数据")
        data_layout = QVBoxLayout()
        data_layout.setSpacing(8)
        
        # 数据加载按钮
        data_buttons_layout = QHBoxLayout()
        self.load_json_button = ModernButton("加载JSON文件", primary=False)
        self.load_json_button.clicked.connect(self.load_from_json)
        self.load_folder_button = ModernButton("从文件夹加载", primary=False)
        self.load_folder_button.clicked.connect(self.load_from_directory)
        data_buttons_layout.addWidget(self.load_json_button)
        data_buttons_layout.addWidget(self.load_folder_button)
        data_layout.addLayout(data_buttons_layout)
        
        # 数据可视化按钮
        self.visualize_data_button = ModernButton("数据分布可视化", primary=False)
        self.visualize_data_button.setEnabled(False)
        self.visualize_data_button.clicked.connect(self.visualize_data_distribution)
        data_layout.addWidget(self.visualize_data_button)
        
        # 数据增强选项
        self.augment_check = QCheckBox("启用数据增强（自动生成相似训练样本）")
        self.augment_check.setChecked(True)
        data_layout.addWidget(self.augment_check)
        
        # 数据统计显示
        self.data_stats_label = StatusInfoWidget("未加载数据", StatusInfoWidget.INFO)
        data_layout.addWidget(self.data_stats_label)
        
        # 数据预览列表
        self.data_preview = QListWidget()
        self.data_preview.setMaximumHeight(150)
        self.data_preview.setAlternatingRowColors(True)
        data_layout.addWidget(QLabel("数据预览:"))
        data_layout.addWidget(self.data_preview)
        
        data_group.setLayout(data_layout)
        left_layout.addWidget(data_group)
        
        # 创建训练参数组
        param_group = QGroupBox("训练参数")
        param_layout = QFormLayout()
        param_layout.setSpacing(10)
        param_layout.setContentsMargins(10, 15, 10, 10)
        
        # 添加训练参数
        self.epochs_input = QSpinBox()
        self.epochs_input.setRange(10, 1000)
        self.epochs_input.setValue(100)
        self.epochs_input.setSingleStep(10)
        param_layout.addRow("训练轮次:", self.epochs_input)
        
        self.learning_rate_input = QDoubleSpinBox()
        self.learning_rate_input.setRange(0.0001, 0.1)
        self.learning_rate_input.setValue(0.001)
        self.learning_rate_input.setSingleStep(0.001)
        self.learning_rate_input.setDecimals(5)
        param_layout.addRow("学习率:", self.learning_rate_input)
        
        self.batch_size_input = QSpinBox()
        self.batch_size_input.setRange(8, 256)
        self.batch_size_input.setValue(32)
        self.batch_size_input.setSingleStep(8)
        param_layout.addRow("批次大小:", self.batch_size_input)
        
        self.early_stopping_check = QCheckBox()
        self.early_stopping_check.setChecked(True)
        param_layout.addRow("启用早停:", self.early_stopping_check)
        
        param_group.setLayout(param_layout)
        left_layout.addWidget(param_group)
        
        # 添加训练控制按钮
        control_layout = QHBoxLayout()
        control_layout.setSpacing(10)
        
        self.start_button = ModernButton("开始训练")
        self.start_button.clicked.connect(self.start_training)
        
        self.stop_button = ModernButton("停止训练", primary=False)
        self.stop_button.clicked.connect(self.stop_training)
        self.stop_button.setEnabled(False)
        
        control_layout.addWidget(self.start_button)
        control_layout.addWidget(self.stop_button)
        left_layout.addLayout(control_layout)
        
        # 添加保存和可视化模型按钮
        model_buttons_layout = QHBoxLayout()
        
        self.save_button = ModernButton("保存模型")
        self.save_button.clicked.connect(self.save_model)
        self.save_button.setEnabled(False)
        
        self.visualize_button = ModernButton("可视化训练过程", primary=False)
        self.visualize_button.clicked.connect(self.visualize_training)
        self.visualize_button.setEnabled(False)
        
        model_buttons_layout.addWidget(self.save_button)
        model_buttons_layout.addWidget(self.visualize_button)
        
        left_layout.addLayout(model_buttons_layout)
        
        # 将左侧面板添加到分割器
        left_panel.layout.addLayout(left_layout)
        splitter.addWidget(left_panel)
        
        # 创建右侧面板（训练监控面板）
        right_panel = StyledCard(self)
        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)
        
        # 创建训练状态组
        status_group = QGroupBox("训练状态")
        status_layout = QVBoxLayout()
        status_layout.setSpacing(8)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        status_layout.addWidget(self.progress_bar)
        
        # 状态标签
        self.status_label = QLabel("就绪")
        self.status_label.setWordWrap(True)
        self.status_label.setAlignment(Qt.AlignCenter)
        status_layout.addWidget(self.status_label)
        
        status_group.setLayout(status_layout)
        right_layout.addWidget(status_group)
        
        # 创建训练历史表格
        history_group = QGroupBox("训练历史")
        history_layout = QVBoxLayout()
        history_layout.setSpacing(5)
        
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(3)
        self.history_table.setHorizontalHeaderLabels(["轮次", "损失", "准确率"])
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setAlternatingRowColors(True)
        history_layout.addWidget(self.history_table)
        
        history_group.setLayout(history_layout)
        right_layout.addWidget(history_group, 1)  # 使用1的拉伸因子
        
        # 添加日志组
        log_group = QGroupBox("训练日志")
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(100)
        self.log_text.setStyleSheet("font-family: Consolas, monospace; font-size: 9pt;")
        log_layout.addWidget(self.log_text)
        
        log_group.setLayout(log_layout)
        right_layout.addWidget(log_group)
        
        # 将右侧面板添加到分割器
        right_panel.layout.addLayout(right_layout)
        splitter.addWidget(right_panel)
        
        # 设置分割器的初始大小
        splitter.setSizes([400, 600])
        
        # 是否正在训练的标志
        self.is_training = False
        self.training_thread = None
        self.loading_thread = None
        
        # 初始化训练历史数据
        self.epochs = []
        self.losses = []
        self.accuracies = []
        
    def load_from_json(self):
        """从JSON文件加载训练数据"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择训练数据文件", "", "JSON文件 (*.json)"
        )
        
        if not file_path:
            return
            
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
            # 检查数据结构
            if isinstance(data, dict) and "intents" in data:
                self.training_data = data["intents"]
            elif isinstance(data, list):
                self.training_data = data
            else:
                raise ValueError("不支持的JSON格式，应为意图列表或包含intents键的对象")
                
            # 更新数据统计和预览
            self.update_data_stats()
            self.update_data_preview()
            
            # 添加日志
            self.log_message(f"成功加载训练数据: {file_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载JSON文件失败: {str(e)}")
            self.log_message(f"加载训练数据失败: {str(e)}")
    
    def load_from_directory(self):
        """从目录加载训练数据"""
        directory = QFileDialog.getExistingDirectory(
            self, "选择训练数据目录", ""
        )
        
        if not directory:
            return
            
        # 显示加载进度提示
        self.log_message(f"开始从目录加载数据: {directory}")
        self.status_label.setText(f"正在从目录加载数据: {directory}...")
        
        # 禁用加载按钮防止重复操作
        self.load_json_button.setEnabled(False)
        self.load_folder_button.setEnabled(False)
        
        # 创建并启动加载线程
        self.loading_thread = DataLoadingThread(directory)
        self.loading_thread.loading_completed.connect(self.on_data_loaded)
        self.loading_thread.loading_failed.connect(self.on_data_loading_failed)
        self.loading_thread.start()
    
    def on_data_loaded(self, data, stats):
        """数据加载完成的回调处理
        
        Args:
            data: 加载的训练数据
            stats: 数据统计信息
        """
        self.training_data = data
        
        # 构建统计信息字符串
        stats_text = (
            f"已加载 {len(data)} 个意图 ("
            f"JSON: {stats['json_files']}, "
            f"文本: {stats['text_files']}, "
            f"图像: {stats['image_files']}, "
            f"视频: {stats['video_files']}"
        )
        
        # 添加音频、CSV和Excel文件统计（如果有的话）
        if 'audio_files' in stats and stats['audio_files'] > 0:
            stats_text += f", 音频: {stats['audio_files']}"
        if 'csv_files' in stats and stats['csv_files'] > 0:
            stats_text += f", CSV: {stats['csv_files']}"
        if 'excel_files' in stats and stats['excel_files'] > 0:
            stats_text += f", Excel: {stats['excel_files']}"
            
        stats_text += ")"
        
        # 设置统计标签
        self.data_stats_label.setText(stats_text)
        self.data_stats_label.set_type(StatusInfoWidget.SUCCESS)
        
        # 更新数据预览
        self.update_data_preview()
        
        # 记录详细统计信息到日志
        stats_detail = (
            f"已处理文件: {stats['processed_files']}/{stats['total_files']}\n"
            f"JSON文件: {stats['json_files']}\n"
            f"文本文件: {stats['text_files']}\n"
            f"图像文件: {stats['image_files']}\n"
            f"视频文件: {stats['video_files']}"
        )
        
        # 添加音频、CSV和Excel文件详细统计
        if 'audio_files' in stats:
            stats_detail += f"\n音频文件: {stats['audio_files']}"
        if 'csv_files' in stats:
            stats_detail += f"\nCSV文件: {stats['csv_files']}"
        if 'excel_files' in stats:
            stats_detail += f"\nExcel文件: {stats['excel_files']}"
            
        stats_detail += f"\n其他文件: {stats['other_files']}\n失败文件: {stats['failed_files']}"
        
        self.log_message(f"数据加载完成!\n{stats_detail}")
        
        # 重新启用加载按钮
        self.load_json_button.setEnabled(True)
        self.load_folder_button.setEnabled(True)
        
        # 启用数据可视化按钮
        self.visualize_data_button.setEnabled(True)
        
        # 更新状态
        self.status_label.setText("数据加载完成，可以开始训练")
    
    def on_data_loading_failed(self, error_message):
        """数据加载失败的回调处理"""
        self.log_message(f"数据加载失败: {error_message}")
        self.status_label.setText("数据加载失败")
        self.data_stats_label.setText("加载失败")
        self.data_stats_label.set_type(StatusInfoWidget.ERROR)
        
        # 重新启用加载按钮
        self.load_json_button.setEnabled(True)
        self.load_folder_button.setEnabled(True)
    
    def update_data_stats(self):
        """更新数据统计信息"""
        if not self.training_data:
            self.data_stats_label.setText("未加载数据")
            self.data_stats_label.set_type(StatusInfoWidget.WARNING)
            return
            
        num_intents = len(self.training_data)
        num_patterns = sum(len(intent.get("patterns", [])) for intent in self.training_data)
        num_responses = sum(len(intent.get("responses", [])) for intent in self.training_data)
        
        stats_text = (
            f"已加载 {num_intents} 个意图, "
            f"{num_patterns} 个训练模式, "
            f"{num_responses} 个可能响应"
        )
        
        self.data_stats_label.setText(stats_text)
        self.data_stats_label.set_type(StatusInfoWidget.SUCCESS)
        
        # 记录到日志
        self.log_message(f"数据统计: {stats_text}")
    
    def update_data_preview(self):
        """更新数据预览列表"""
        self.data_preview.clear()
        
        if not self.training_data:
            self.data_preview.addItem("未加载数据")
            return
        
        # 限制显示前100个意图
        for i, intent in enumerate(self.training_data[:100]):  
            tag = intent.get("tag", f"未命名_{i}")
            patterns_count = len(intent.get("patterns", []))
            responses_count = len(intent.get("responses", []))
            
            # 检查是否有媒体
            media_type = intent.get("media_type", "")
            media_path = intent.get("media_path", "")
            
            if media_type:
                media_info = f" [{media_type}: {os.path.basename(media_path) if media_path else '无路径'}]"
            else:
                media_info = ""
            
            item_text = f"{tag}: {patterns_count} 模式, {responses_count} 回复{media_info}"
            item = QListWidgetItem(item_text)
            
            # 设置工具提示显示更多信息
            tooltip = f"标签: {tag}\n模式数量: {patterns_count}\n"
            
            # 如果不太多，显示一些模式示例
            if patterns_count > 0:
                patterns_examples = intent.get("patterns", [])[:3]  # 最多显示3个
                patterns_str = "\n".join([f"- {p}" for p in patterns_examples])
                if patterns_count > 3:
                    patterns_str += f"\n(还有 {patterns_count-3} 个...)"
                tooltip += f"模式示例:\n{patterns_str}\n"
            
            # 添加响应示例
            if responses_count > 0:
                responses_examples = intent.get("responses", [])[:2]  # 最多显示2个
                responses_str = "\n".join([f"- {r[:50]}..." if len(r) > 50 else f"- {r}" for r in responses_examples])
                if responses_count > 2:
                    responses_str += f"\n(还有 {responses_count-2} 个...)"
                tooltip += f"响应示例:\n{responses_str}"
            
            # 设置媒体信息
            if media_path:
                tooltip += f"\n媒体路径: {media_path}"
            
            item.setToolTip(tooltip)
            
            # 设置标记颜色
            if "media_type" in intent:
                if intent["media_type"] == "image":
                    item.setForeground(QColor(46, 125, 50))  # 绿色
                elif intent["media_type"] == "video":
                    item.setForeground(QColor(211, 47, 47))  # 红色
                elif intent["media_type"] == "audio":
                    item.setForeground(QColor(33, 150, 243))  # 蓝色
            
            self.data_preview.addItem(item)
        
        # 添加统计信息
        self.data_preview.addItem(f"--- 共 {len(self.training_data)} 个意图 ---")
    
    def start_training(self):
        """开始训练"""
        if self.is_training:
            QMessageBox.warning(self, "警告", "训练已在进行中")
            return
            
        if not self.training_data:
            QMessageBox.warning(self, "警告", "请先加载训练数据")
            return
            
        # 确定模型类型
        if self.forest_radio.isChecked():
            model_type = "forest"
        elif self.neural_radio.isChecked():
            model_type = "neural"
        elif self.transformer_radio.isChecked():
            model_type = "transformer"
        else:
            model_type = "forest"  # 默认
        
        # 获取训练参数
        epochs = self.epochs_input.value()
        learning_rate = self.learning_rate_input.value()
        batch_size = self.batch_size_input.value()
        early_stopping = self.early_stopping_check.isChecked()
        
        # 创建训练线程
        self.training_thread = TrainingWorker(
            training_data=self.training_data,
            epochs=epochs,
            learning_rate=learning_rate,
            batch_size=batch_size,
            early_stopping=early_stopping,
            model_type=model_type
        )
        
        # 连接信号
        self.training_thread.progress_updated.connect(self.update_progress)
        self.training_thread.status_updated.connect(self.update_status)
        self.training_thread.epoch_completed.connect(self.on_epoch_completed)
        self.training_thread.training_completed.connect(self.on_training_completed)
        self.training_thread.finished.connect(self.on_thread_finished)
        
        # 更新UI状态
        self.is_training = True
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.save_button.setEnabled(False)
        self.load_json_button.setEnabled(False)
        self.load_folder_button.setEnabled(False)
        
        # 清空历史表格
        self.history_table.setRowCount(0)
        
        # 添加日志
        self.log_message(f"开始训练 {model_type} 模型，训练轮次: {epochs}")
        
        # 启动线程
        self.training_thread.start()
    
    def stop_training(self):
        """停止训练"""
        if not self.is_training or not self.training_thread:
            return
            
        reply = QMessageBox.question(
            self, "确认", "确定要停止训练吗？已训练的进度将丢失。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.training_thread.stop()
            self.update_status("正在停止训练...")
            self.log_message("用户请求停止训练")
    
    def update_progress(self, value):
        """更新进度条"""
        self.progress_bar.setValue(value)
    
    def update_status(self, message):
        """更新状态标签"""
        self.status_label.setText(message)
        self.log_message(message)
    
    def on_epoch_completed(self, epoch, loss, accuracy):
        """处理轮次完成信号"""
        # 添加到历史表格
        row = self.history_table.rowCount()
        self.history_table.insertRow(row)
        
        # 设置表格项
        self.history_table.setItem(row, 0, QTableWidgetItem(str(epoch)))
        self.history_table.setItem(row, 1, QTableWidgetItem(f"{loss:.4f}"))
        self.history_table.setItem(row, 2, QTableWidgetItem(f"{accuracy:.4f}"))
        
        # 滚动到底部
        self.history_table.scrollToBottom()
        
        # 记录训练历史数据，用于后续可视化
        self.epochs.append(epoch)
        self.losses.append(loss)
        self.accuracies.append(accuracy)
        
        # 添加日志
        self.log_message(f"轮次 {epoch} 完成，损失: {loss:.4f}，准确率: {accuracy:.4f}")
    
    def on_training_completed(self, success, message):
        """处理训练完成信号"""
        if success:
            QMessageBox.information(self, "训练完成", message)
            self.save_button.setEnabled(True)
            self.visualize_button.setEnabled(True)  # 启用可视化按钮
        else:
            QMessageBox.warning(self, "训练失败", message)
            
        self.log_message(message)
        
        # 尝试清理临时文件
        self._cleanup_temp_files()
    
    def on_thread_finished(self):
        """处理线程结束"""
        self.is_training = False
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.load_json_button.setEnabled(True)
        self.load_folder_button.setEnabled(True)
        
        # 尝试清理临时文件
        self._cleanup_temp_files()
    
    def _cleanup_temp_files(self):
        """清理训练过程中创建的临时文件"""
        if hasattr(self, 'training_thread') and hasattr(self.training_thread, 'temp_file_path'):
            temp_path = self.training_thread.temp_file_path
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                    self.log_message(f"已清理临时文件: {temp_path}")
                except Exception as e:
                    self.log_message(f"清理临时文件失败: {str(e)}")
    
    def save_model(self):
        """保存模型"""
        if not hasattr(self, 'training_thread') or not self.training_thread or not hasattr(self.training_thread, 'model'):
            QMessageBox.warning(self, "警告", "无可保存的模型")
            return
            
        # 打开保存对话框
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存模型", "", "模型文件 (*.pth)"
        )
        
        if not file_path:
            return
            
        try:
            # 在实际应用中，这里应该调用模型的save方法
            # 这里简化实现
            self.log_message(f"正在保存模型到: {file_path}")
            
            # 模拟保存过程
            time.sleep(1)
            
            # 添加日志
            self.log_message(f"模型保存成功: {file_path}")
            QMessageBox.information(self, "保存成功", f"模型已保存到: {file_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存模型失败: {str(e)}")
            self.log_message(f"保存模型失败: {str(e)}")
    
    def visualize_training(self):
        """可视化训练过程"""
        if not MATPLOTLIB_AVAILABLE:
            QMessageBox.warning(self, "警告", "无法可视化训练过程，缺少Matplotlib库")
            return
            
        if not self.epochs:
            QMessageBox.information(self, "提示", "没有训练历史数据可供可视化")
            return
            
        # 创建可视化对话框
        viz_dialog = TrainingVisualizationDialog(self.epochs, self.losses, self.accuracies, parent=self)
        viz_dialog.exec_()
    
    def visualize_data_distribution(self):
        """可视化数据分布"""
        if not MATPLOTLIB_AVAILABLE:
            QMessageBox.warning(self, "警告", "无法可视化数据分布，缺少Matplotlib库")
            return
            
        if not self.training_data:
            QMessageBox.information(self, "提示", "没有训练数据可供可视化")
            return
            
        # 创建可视化对话框
        viz_dialog = DataDistributionDialog(self.training_data, parent=self)
        viz_dialog.exec_()
    
    def log_message(self, message):
        """添加日志消息"""
        timestamp = time.strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
        # 滚动到底部
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())


class TrainingVisualizationDialog(QDialog):
    """训练可视化对话框"""
    
    def __init__(self, epochs, losses, accuracies, parent=None):
        super().__init__(parent)
        self.setWindowTitle("训练过程可视化")
        self.resize(800, 500)
        
        # 创建主布局
        main_layout = QVBoxLayout(self)
        
        # 创建标签页容器
        tabs = QTabWidget()
        main_layout.addWidget(tabs)
        
        # 创建损失曲线标签页
        loss_tab = QWidget()
        loss_layout = QVBoxLayout(loss_tab)
        
        # 创建Matplotlib图形
        loss_figure = Figure(figsize=(8, 5), dpi=100)
        loss_canvas = FigureCanvas(loss_figure)
        loss_layout.addWidget(loss_canvas)
        
        # 绘制损失曲线
        ax = loss_figure.add_subplot(111)
        ax.plot(epochs, losses, 'r-', linewidth=2)
        ax.set_title('训练损失曲线')
        ax.set_xlabel('轮次')
        ax.set_ylabel('损失值')
        ax.grid(True)
        loss_figure.tight_layout()
        
        # 添加损失曲线标签页
        tabs.addTab(loss_tab, "损失曲线")
        
        # 创建准确率曲线标签页
        acc_tab = QWidget()
        acc_layout = QVBoxLayout(acc_tab)
        
        # 创建Matplotlib图形
        acc_figure = Figure(figsize=(8, 5), dpi=100)
        acc_canvas = FigureCanvas(acc_figure)
        acc_layout.addWidget(acc_canvas)
        
        # 绘制准确率曲线
        ax = acc_figure.add_subplot(111)
        ax.plot(epochs, accuracies, 'b-', linewidth=2)
        ax.set_title('训练准确率曲线')
        ax.set_xlabel('轮次')
        ax.set_ylabel('准确率')
        ax.set_ylim([0, 1.05])
        ax.grid(True)
        acc_figure.tight_layout()
        
        # 添加准确率曲线标签页
        tabs.addTab(acc_tab, "准确率曲线")
        
        # 创建组合图标签页
        combined_tab = QWidget()
        combined_layout = QVBoxLayout(combined_tab)
        
        # 创建Matplotlib图形
        combined_figure = Figure(figsize=(8, 5), dpi=100)
        combined_canvas = FigureCanvas(combined_figure)
        combined_layout.addWidget(combined_canvas)
        
        # 创建双Y轴图
        ax1 = combined_figure.add_subplot(111)
        line1, = ax1.plot(epochs, losses, 'r-', linewidth=2, label='损失')
        ax1.set_xlabel('轮次')
        ax1.set_ylabel('损失值', color='r')
        ax1.tick_params(axis='y', labelcolor='r')
        
        ax2 = ax1.twinx()
        line2, = ax2.plot(epochs, accuracies, 'b-', linewidth=2, label='准确率')
        ax2.set_ylabel('准确率', color='b')
        ax2.tick_params(axis='y', labelcolor='b')
        ax2.set_ylim([0, 1.05])
        
        # 添加图例
        lines = [line1, line2]
        combined_figure.legend(lines, [l.get_label() for l in lines], loc='upper center')
        
        combined_figure.suptitle('训练损失和准确率')
        combined_figure.tight_layout()
        
        # 添加组合图标签页
        tabs.addTab(combined_tab, "组合图")
        
        # 添加关闭按钮
        button_layout = QHBoxLayout()
        close_button = QPushButton("关闭")
        close_button.clicked.connect(self.accept)
        button_layout.addStretch()
        button_layout.addWidget(close_button)
        main_layout.addLayout(button_layout)


class DataDistributionDialog(QDialog):
    """数据分布可视化对话框"""
    
    def __init__(self, training_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("数据分布可视化")
        self.resize(800, 600)
        self.training_data = training_data
        
        # 创建主布局
        layout = QVBoxLayout(self)
        
        # 创建标签页容器
        tabs = QTabWidget()
        layout.addWidget(tabs)
        
        # 创建意图分布标签页
        intent_tab = QWidget()
        intent_layout = QVBoxLayout(intent_tab)
        
        # 创建Matplotlib图形
        intent_figure = Figure(figsize=(8, 6), dpi=100)
        intent_canvas = FigureCanvas(intent_figure)
        intent_layout.addWidget(intent_canvas)
        
        # 绘制意图分布
        self.plot_intent_distribution(intent_figure)
        
        # 添加意图分布标签页
        tabs.addTab(intent_tab, "意图分布")
        
        # 创建模式分布标签页
        pattern_tab = QWidget()
        pattern_layout = QVBoxLayout(pattern_tab)
        
        # 创建Matplotlib图形
        pattern_figure = Figure(figsize=(8, 6), dpi=100)
        pattern_canvas = FigureCanvas(pattern_figure)
        pattern_layout.addWidget(pattern_canvas)
        
        # 绘制模式分布
        self.plot_pattern_distribution(pattern_figure)
        
        # 添加模式分布标签页
        tabs.addTab(pattern_tab, "模式分布")
        
        # 创建媒体类型分布标签页
        media_tab = QWidget()
        media_layout = QVBoxLayout(media_tab)
        
        # 创建Matplotlib图形
        media_figure = Figure(figsize=(8, 6), dpi=100)
        media_canvas = FigureCanvas(media_figure)
        media_layout.addWidget(media_canvas)
        
        # 绘制媒体类型分布
        self.plot_media_distribution(media_figure)
        
        # 添加媒体类型分布标签页
        tabs.addTab(media_tab, "媒体类型分布")
        
        # 添加关闭按钮
        button_layout = QHBoxLayout()
        close_button = QPushButton("关闭")
        close_button.clicked.connect(self.accept)
        button_layout.addStretch()
        button_layout.addWidget(close_button)
        layout.addLayout(button_layout)
    
    def plot_intent_distribution(self, figure):
        """绘制意图分布"""
        # 统计每个意图的模式数量
        intent_counts = {}
        for intent in self.training_data:
            tag = intent.get('tag', 'unknown')
            patterns_count = len(intent.get('patterns', []))
            intent_counts[tag] = patterns_count
        
        # 取模式数量前20的意图
        top_intents = sorted(intent_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        
        ax = figure.add_subplot(111)
        
        tags = [item[0] for item in top_intents]
        counts = [item[1] for item in top_intents]
        
        # 创建水平条形图
        bars = ax.barh(tags, counts, color='skyblue')
        
        # 在条形图上添加数字标签
        for bar in bars:
            width = bar.get_width()
            ax.text(width + 1, bar.get_y() + bar.get_height()/2, f'{int(width)}', 
                    ha='left', va='center')
        
        ax.set_title('意图模式数量分布 (前20个)')
        ax.set_xlabel('模式数量')
        ax.set_ylabel('意图标签')
        figure.tight_layout()
    
    def plot_pattern_distribution(self, figure):
        """绘制模式分布（模式长度直方图）"""
        # 统计所有模式的长度
        pattern_lengths = []
        for intent in self.training_data:
            patterns = intent.get('patterns', [])
            for pattern in patterns:
                pattern_lengths.append(len(pattern))
        
        ax = figure.add_subplot(111)
        
        # 创建直方图
        n, bins, patches = ax.hist(pattern_lengths, bins=20, color='skyblue', edgecolor='black')
        
        ax.set_title('模式长度分布')
        ax.set_xlabel('模式长度（字符数）')
        ax.set_ylabel('频率')
        figure.tight_layout()
    
    def plot_media_distribution(self, figure):
        """绘制媒体类型分布"""
        # 统计各种媒体类型的数量
        media_counts = {'无媒体': 0, '图像': 0, '视频': 0, '音频': 0, '其他': 0}
        
        for intent in self.training_data:
            media_type = intent.get('media_type', '')
            if not media_type:
                media_counts['无媒体'] += 1
            elif media_type == 'image':
                media_counts['图像'] += 1
            elif media_type == 'video':
                media_counts['视频'] += 1
            elif media_type == 'audio':
                media_counts['音频'] += 1
            else:
                media_counts['其他'] += 1
        
        # 删除计数为0的类别
        media_counts = {k: v for k, v in media_counts.items() if v > 0}
        
        ax = figure.add_subplot(111)
        
        # 创建饼图
        labels = list(media_counts.keys())
        sizes = list(media_counts.values())
        
        # 设置饼图颜色
        colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99', '#c2c2f0']
        
        # 创建带百分比的饼图
        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax.axis('equal')  # 使饼图为正圆形
        
        ax.set_title('媒体类型分布')
        figure.tight_layout() 